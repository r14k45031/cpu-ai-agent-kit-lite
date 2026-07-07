# -*- coding: utf-8 -*-
"""Excel AI 欄位處理：AI 逐列處理資料，把結果寫進新欄位。

例：拖入客訴清單，指示「根據備註欄判斷急件或一般件」→ 每列多一欄分類結果。

用法：excel_process.py <xlsx/csv 路徑>（執行時會詢問任務指示與新欄位名稱）
輸出：<原檔名>_AI處理.xlsx（原檔不動）
"""
import csv
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
import ai_common as ai

BATCH = 5


def load_rows(f):
    """回傳 (header, rows)。csv 也轉成同樣結構。"""
    if f.suffix.lower() == ".csv":
        with open(f, newline="", encoding="utf-8-sig", errors="ignore") as fh:
            data = list(csv.reader(fh))
        if not data:
            raise SystemExit("檔案是空的。")
        return data[0], data[1:], None
    import openpyxl

    wb = openpyxl.load_workbook(str(f))
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise SystemExit("第一個工作表是空的。")
    print(f"處理工作表：{ws.title}（其餘工作表原樣保留）")
    return list(rows[0]), [list(r) for r in rows[1:]], (wb, ws)


def row_text(header, row):
    pairs = []
    for k, v in zip(header, row):
        if v is None or str(v).strip() == "":
            continue
        pairs.append(f"{k if k else '欄位'}：{v}")
    return "；".join(pairs)


def process_batch(texts, instr, model):
    """一次送 BATCH 列，回傳結果清單；解析失敗回 None。"""
    numbered = "\n".join(f"{i}. {t}" for i, t in enumerate(texts, 1))
    ans = ai.chat(
        f"以下是表格中的 {len(texts)} 列資料：\n{numbered}\n\n"
        f"任務：對每一列，{instr}\n"
        f"輸出格式：共 {len(texts)} 行，每行「編號. 結果」，"
        "結果要精簡（一個詞或一句話），不要任何其他文字。"
        "一律使用繁體中文（台灣用語）。",
        model=model,
    )
    results = {}
    for line in ans.splitlines():
        m = re.match(r"\s*(\d+)[\.、:：]\s*(.+)", line.strip())
        if m and 1 <= int(m.group(1)) <= len(texts):
            results[int(m.group(1))] = m.group(2).strip()
    if len(results) != len(texts):
        return None
    return [results[i] for i in range(1, len(texts) + 1)]


def process_one(text, instr, model):
    return ai.chat(
        f"表格中的一列資料：{text}\n任務：{instr}\n"
        "只輸出結果（一個詞或一句話），不要任何說明。繁體中文。",
        model=model,
    ).strip()


def main():
    if len(sys.argv) < 2:
        raise SystemExit("用法：把 .xlsx / .csv 拖到「拖入Excel欄位AI處理.bat」上")
    f = Path(sys.argv[1])
    if f.suffix.lower() not in (".xlsx", ".xlsm", ".csv"):
        raise SystemExit("請拖入 .xlsx / .xlsm / .csv 檔案")

    header, rows, wb_ws = load_rows(f)
    print(f"欄位：{'、'.join(str(h) for h in header if h)}（共 {len(rows)} 列資料）")
    try:
        instr = input("要對每一列做什麼？（例：根據備註欄判斷急件或一般件）\n> ").strip()
    except EOFError:
        instr = ""
    if not instr:
        raise SystemExit("需要輸入任務指示才能處理。")
    try:
        col_name = input("新欄位名稱？（直接按 Enter ＝ AI結果）\n> ").strip() or "AI結果"
    except EOFError:
        col_name = "AI結果"

    model = ai.pick_model()
    if len(rows) > 300:
        print(f"注意：共 {len(rows)} 列，純 CPU 可能需要很久（估每列 3~10 秒），可隨時 Ctrl+C 中斷。")
    print(f"開始處理（模型：{model}，每 {BATCH} 列一批）...")

    texts = [row_text(header, r) for r in rows]
    results = []
    for start in range(0, len(texts), BATCH):
        chunk = texts[start:start + BATCH]
        batch_res = process_batch(chunk, instr, model) if len(chunk) > 1 else None
        if batch_res is None:
            batch_res = [process_one(t, instr, model) if t else "" for t in chunk]
        results.extend(batch_res)
        print(f"  已完成 {min(start + BATCH, len(texts))}/{len(texts)} 列")

    out = f.with_name(f.stem + "_AI處理.xlsx")
    if wb_ws:
        wb, ws = wb_ws
        col = ws.max_column + 1
        ws.cell(row=1, column=col, value=col_name)
        for i, val in enumerate(results, 2):
            ws.cell(row=i, column=col, value=val)
        wb.save(str(out))
    else:
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(list(header) + [col_name])
        for r, val in zip(rows, results):
            ws.append(list(r) + [val])
        wb.save(str(out))
    print(f"完成：已輸出 {out.name}（新欄位「{col_name}」）")
    print("提醒：xlsx 內的圖表與圖片不會帶到新檔，純資料表不受影響。")


if __name__ == "__main__":
    main()
